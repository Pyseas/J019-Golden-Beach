
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import numpy as np

PATH = Path('.')


def write_piping_loads(xlname: str, outname: str) -> None:
    """Write load data from a spreadsheet to a SACS format file.

    Args:
        xlname (str): Spreadsheet filename.
        outname (str): Output filename.

    """

    xlpath = PATH.joinpath(xlname)
    outpath = PATH.joinpath(outname)

    loadcns = pd.read_excel(xlpath, sheet_name='Load Case ID', converters={'SUFFIX': str})

    outstr = ''
    for row in loadcns.itertuples():
        sheet = row.Sheet
        col = int(row.Column)
        loadcn = row.LOADCN
        suffix = row.SUFFIX
        loadlb = row.LOADLB
        loadid = row.LOAD_ID

        wb = load_workbook(xlpath, data_only=True)
        ws = wb[sheet]
        sup_labels = [ws.cell(row=i, column=1).value for i in range(3, 22)]
        data = []
        for irow in range(3, 22):
            data.append([ws.cell(row=irow, column=i).value for i in range(col, col + 6)])
        data = np.asarray(data) / 1000
        # np.savetxt(PATH.joinpath('test.txt'), data)

        outstr += 'LOADCN' + f'{loadcn: >4} 1.00\n'
        outstr += 'LOADLB' + f'{loadcn: >4} {loadlb}\n'

        ndata = np.shape(data)[0]
        for irow in range(ndata):
            joint = sup_labels[irow]
            loads = data[irow, :]
            outstr += f'LOAD{joint: >7}' + ' '*5
            for val in loads[:4]:
                outstr += f'{val:7.1f}'
            outstr += ' '
            for val in loads[4:]:
                outstr += f'{val:7.1f}'
            if loadid == 'PSXX':
                remark = f'{joint}_{suffix}'
            else:
                remark = loadid
            outstr += ' GLOB JOIN   '
            outstr += f'{remark: >8}\n'

    with open(outpath, 'w') as f:
        f.write(outstr)


def main():

    xlname = 'connector_loads_20241003.xlsx'
    outname = 'loadcn.txt'

    write_piping_loads(xlname, outname)


if __name__ == "__main__":
    main()
